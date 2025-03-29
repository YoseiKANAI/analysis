# %%
# # -*- coding: utf-8 -*-
"""
Created on: 2025-01-04 20:35

@author: ShimaLab
"""

import os
import glob
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import statistics as st
from scipy.signal import welch
import openpyxl

import global_value as g

Fs = 2000
task_num = len(g.task)

# FFTの設定
window_length = 5.0  #　ウィンドウの幅[s]
T = 30     # Total duration in seconds
# Define window parameters for analysis
window_size = int(Fs * window_length)  # Window size in samples
num_windows = int(T // window_length)

def main():
    output_name, output_dir_plot = output_preparation()
    wb = openpyxl.Workbook()
    MNF_summary = pd.DataFrame()
    for ID in range(g.subnum):
        # パスを指定
        input_dir = "D:/User/kanai/Data/%s/sub%s/csv/*_a_2.csv" % (g.datafile, ID+1)
        output_dir_indi = output_preparation_indi(ID)
        # 各試行全体のMNFを計算
        df = MNF(wb, input_dir, output_name, ID, output_dir_indi)
        #sub = "sub%d" %(ID+1)
        #MNF_summary = pd.concat([MNF_summary, pd.DataFrame([sub])])
        #MNF_summary = pd.concat([MNF_summary, df])
    #sheet_name = "MNF_whole"
    #excel_output(MNF_summary, output_name, sheet_name)
        
"""
MNFのメインの関数
"""
def MNF(wb, input_dir, output_name, ID, output_dir_indi):
    task_list = []
    sheet_name = "MNF_sub%d" %(ID+1)
    filelist = glob.glob(input_dir)
    
    mean = []
    std = []
    
    # リストの順に呼び出し
    for t in g.task:
        cnt = 0
        task_list = [s for s in filelist if t in s]
        num_windows = int(T // window_length)
        #t_data = np.empty((g.attempt, num_windows, g.muscle_num))
        t_data = []

        for file in task_list:
            attempt = file[35:41]
            task_name = file[35:37]
            df , columns_list = csv_reader(file)
            output_file_indi = output_dir_indi + attempt + ".csv" 

            """
            if (attempt =="FB0001" or attempt =="DW0004") and ID ==6:
                MNF = pd.concat([MNF, pd.DataFrame(index=[task_name])])
                continue
            if attempt =="FB0003" and ID ==9:
                MNF = pd.concat([MNF, pd.DataFrame(index=[task_name])])
                continue
            
            # データがうまく取れてないやつ
            if attempt =="NC0002" and ID ==2:
                MNF = pd.concat([MNF, pd.DataFrame(index=[task_name])])
                continue            
            """
            
            MNF, time_points = MNF_cal(df.interpolate("index"))
            MNF = MNF.iloc[:, 0:12]
            MNF.columns = g.muscle_columns
            MNF.to_csv(output_file_indi)
            
            #MNF_norm = MNF.divide(MNF.iloc[0], axis=1)
            #なんか格納する配列を作成して，平均をとる
            t_data.append(MNF.values)
        data = np.array(t_data)
        divine = np.nanmean(data[:, 0, :], axis = 0)
        data_norm = data/divine
        mean.append(np.nanmean(data, axis=0))
        std.append(np.nanstd(data, axis=0, ddof=1))
    
    output_excel_wb(wb, mean, std, output_name, sheet_name)

"""
csvの読み込みを行う関数
"""
def csv_reader(file):
    df = pd.read_csv(file)
    columns_list = df.columns.values
    return df, columns_list

"""
5sごとのMNFを算出
"""
def MNF_cal(df):
    mean_freqs = pd.DataFrame()
    time_points = []
    # Loop over each window to compute the mean frequency
    for i in range(num_windows):
        start_idx = i * window_size
        end_idx = start_idx + window_size
        segment = df[start_idx:end_idx]
        
        # Compute the Power Spectral Density (PSD) using Welch's method
        nperseg = 2048  # Length of each segment for Welch's method
        freqs, Pxx = welch(segment, fs=Fs, nperseg=nperseg, axis = 0)
        
        a = np.sum(Pxx*freqs[:, np.newaxis], axis=0) 
        b = np.sum(Pxx, axis=0)
        
        # Compute the mean frequency
        mean_freq = pd.DataFrame(a/b).T
        mean_freqs = pd.concat([mean_freqs, mean_freq], axis = 0)
        
        # Calculate the time point (center of the window)
        time_point = (start_idx + end_idx) / 2 / Fs
        time_points.append(time_point)
    return mean_freqs, time_points

"""
MNFを計算, シートに書き込み

def MNF_cal(psd, freq, columns_list):
    MNF = pd.DataFrame([np.zeros(g.muscle_num)], columns = columns_list)
    sum_numr = pd.DataFrame([np.zeros(g.muscle_num)], columns = columns_list)
    sum_deno = pd.DataFrame([np.zeros(g.muscle_num)], columns = columns_list)
    
    for i in range(len(psd)):
        sum_numr += psd[i, :]*freq[i]
        sum_deno += psd[i, :]
        
    MNF = sum_numr / sum_deno
    return MNF
"""

"""
excelに出力
"""
def excel_output(data, output_name, sheet_name):
    if (os.path.isfile(output_name)):
        with pd.ExcelWriter(output_name, mode="a") as writer:
            data.to_excel(writer, sheet_name = sheet_name)
    else:
        with pd.ExcelWriter(output_name) as writer:
            data.to_excel(writer, sheet_name = sheet_name)        

"""
openpyxlを用いてexcel出力
"""
def output_excel_wb(wb, mean, std, output_name, sheet_name):
    # ワークブックとシートの作成
    ws = wb.create_sheet(title=sheet_name)

    for t in range(task_num):
        # list1を書き込む
        for row_idx, row in enumerate(mean[t], start=(len(mean[0])+1)*t+1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # list2を書き込む (list1の右に空白を挟む)
        start_col = len(mean[0][0]) + 2  # list1の列数 + 空白の幅
        for row_idx, row in enumerate(std[t], start=(len(mean[0])+1)*t+1):
            for col_idx, value in enumerate(row, start=start_col):
                ws.cell(row=row_idx, column=col_idx, value=value)
    # Excelファイルとして保存
    wb.save(output_name)


"""
ファイル名の定義
"""        
def output_preparation():
    # ルートフォルダのパスを指定
    output_dir = "D:/User/kanai/Data/%s/result_EMG/MNF" %(g.datafile)
    output_dir_plot = "D:/User/kanai/Data/%s/result_EMG/MNF/plot" %(g.datafile)
    output_name = output_dir + "/result.xlsx"
    # 出力先フォルダを作成
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(output_dir_plot, exist_ok=True)
    # エクセルファイルの初期化
    if (os.path.isfile(output_name)):
        os.remove(output_name)
    
    return output_name, output_dir_plot

"""
個別ファイルの格納先指定
"""  
def output_preparation_indi(ID):
    output_dir = "D:/User/kanai/Data/%s/sub%d/csv/MNF/" %(g.datafile, ID+1)
    os.makedirs(output_dir, exist_ok=True)
    
    return output_dir

if __name__ == "__main__":
    main()